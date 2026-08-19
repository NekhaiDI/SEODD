# -*- coding: utf-8 -*-
"""
Генерирует HTML- и Excel-отчёты по результатам сравнения сайтов.
"""
import os
import html as html_lib
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _fmt(v):
    if v is None:
        return "—"
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def _best_index(row):
    """Возвращает индекс сайта с наилучшим значением метрики (или None)."""
    direction = row["direction"]
    values = row["values"]
    if direction == "neutral":
        return None
    numeric = [(i, v) for i, v in enumerate(values) if isinstance(v, (int, float))]
    if not numeric:
        return None
    if direction == "lower_better":
        return min(numeric, key=lambda x: x[1])[0]
    else:
        return max(numeric, key=lambda x: x[1])[0]


def collect_problem_pages(pages_data, limit=100):
    """Страницы с базовыми SEO-проблемами + человекочитаемый список проблем."""
    out = []
    for p in pages_data:
        problems = []
        if not p["has_title"]:
            problems.append("нет title")
        if not p["has_meta_description"]:
            problems.append("нет meta description")
        if not p["has_h1"]:
            problems.append("нет H1")
        if p["multiple_h1"]:
            problems.append("несколько H1")
        if p["is_thin_content"]:
            problems.append(f"мало текста ({p['word_count']} слов)")
        if problems:
            out.append({"url": p["url"], "problems": problems})
        if len(out) >= limit:
            break
    return out


def generate_html_report(comparison, my_site_pages_data, out_path):
    sites = comparison["sites"]
    rows = comparison["rows"]

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    head_cells = "".join(f"<th>{html_lib.escape(s['name'])}</th>" for s in sites)

    body_rows = []
    for row in rows:
        best_idx = _best_index(row)
        cells = []
        for i, v in enumerate(row["values"]):
            cls = " class=\"best\"" if best_idx == i else ""
            cells.append(f"<td{cls}>{html_lib.escape(_fmt(v))}</td>")
        body_rows.append(
            f"<tr><td class=\"metric-name\">{html_lib.escape(row['label'])}</td>{''.join(cells)}</tr>"
        )

    # Список проблемных страниц по своему сайту (первые из выборки)
    problem_pages = collect_problem_pages(my_site_pages_data)
    issues_html = ""
    if problem_pages:
        issue_rows = []
        for p in problem_pages:
            issue_rows.append(
                f"<tr><td><a href=\"{html_lib.escape(p['url'])}\" target=\"_blank\">{html_lib.escape(p['url'])}</a></td>"
                f"<td>{html_lib.escape(', '.join(p['problems']))}</td></tr>"
            )
        issues_html = f"""
        <h2>Страницы твоего сайта с проблемами SEO (первые {len(problem_pages)})</h2>
        <table class="issues">
          <tr><th>URL</th><th>Проблемы</th></tr>
          {''.join(issue_rows)}
        </table>
        """

    pages_row = next(r["values"] for r in rows if r["key"] == "pages_crawled")
    crawled_line = ", ".join(f"{s['name']}: {v}" for s, v in zip(sites, pages_row))

    html_content = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>SEO-отчёт по конкурентам — {timestamp}</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif; margin: 40px; color: #1a1a1a; background: #fafafa; }}
  h1 {{ font-size: 24px; margin-bottom: 4px; }}
  .subtitle {{ color: #666; margin-bottom: 32px; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 40px; background: #fff; }}
  th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: left; font-size: 14px; }}
  th {{ background: #005f86; color: #fff; }}
  td.metric-name {{ font-weight: 600; background: #f0f4f7; }}
  td.best {{ background: #d9f2d9; font-weight: 600; }}
  table.issues td {{ font-size: 13px; }}
  table.issues a {{ color: #005f86; }}
  .footer {{ color: #999; font-size: 12px; margin-top: 40px; }}
</style>
</head>
<body>
  <h1>SEO-сравнение сайта с конкурентами</h1>
  <div class="subtitle">Сформировано: {timestamp} · Просканировано страниц: {crawled_line}</div>

  <table>
    <tr><th>Метрика</th>{head_cells}</tr>
    {''.join(body_rows)}
  </table>

  {issues_html}

  <div class="footer">Зелёным выделено лучшее значение метрики среди сравниваемых сайтов (там, где применимо).</div>
</body>
</html>
"""

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    return out_path


def generate_excel_report(comparison, all_pages_data_by_site, out_path):
    """
    all_pages_data_by_site: dict {site_name: [page_metric_dict, ...]}
    """
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="005F86", end_color="005F86", fill_type="solid")
    best_fill = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")

    # --- Лист "Сравнение" ---
    ws = wb.active
    ws.title = "Сравнение"
    sites = comparison["sites"]
    ws.cell(row=1, column=1, value="Метрика").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for i, s in enumerate(sites):
        c = ws.cell(row=1, column=2 + i, value=s["name"])
        c.font = header_font
        c.fill = header_fill

    for r_idx, row in enumerate(comparison["rows"], start=2):
        ws.cell(row=r_idx, column=1, value=row["label"])
        best_idx = _best_index(row)
        for i, v in enumerate(row["values"]):
            cell = ws.cell(row=r_idx, column=2 + i, value=v if v is not None else "—")
            if best_idx == i:
                cell.fill = best_fill

    for col in range(1, 2 + len(sites)):
        ws.column_dimensions[get_column_letter(col)].width = 32 if col == 1 else 20

    # --- Лист по каждому сайту с постраничными данными ---
    for site_name, pages in all_pages_data_by_site.items():
        sheet_name = site_name[:31]  # ограничение Excel
        ws2 = wb.create_sheet(sheet_name)
        columns = [
            "url", "status_code", "title", "title_length", "has_title",
            "meta_description_length", "has_meta_description",
            "h1_count", "h1_text", "word_count", "is_thin_content",
            "has_canonical", "images_missing_alt", "internal_links",
            "external_links", "response_time_ms",
        ]
        for c_idx, col in enumerate(columns, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill

        for r_idx, page in enumerate(pages, start=2):
            for c_idx, col in enumerate(columns, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=page.get(col))

        ws2.freeze_panes = "A2"
        for col_idx in range(1, len(columns) + 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 22

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
